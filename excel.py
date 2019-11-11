import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("extras/transactions.xlsx")
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)
# print(cell.value)

for row in range(2, sheet.max_row+1): #ignoring 1st row as it is a header
    cell=sheet.cell(row, 3) #we are only interested in 3rd coloum
    new_value = cell.value * 0.9
    new_cell = sheet.cell(row, 4)
    new_cell.value = new_value

bar_ref = Reference(sheet, 
    min_row=2, 
    max_row=sheet.max_row,
    min_col=4,
    max_col=4
    )

chart = BarChart()
chart.add_data(bar_ref)
sheet.add_chart(chart, 'e2')

wb.save('transactions-corrected.xlsx')
